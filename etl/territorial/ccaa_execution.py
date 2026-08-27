"""Normalize the official CIMCANET monthly CCAA execution workbook."""

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


MONTHS = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
SOURCE_URL = "https://serviciostelematicosext.hacienda.gob.es/SGCIEF/Cimcanet/aspx/consulta/consulta.aspx"


def normalize_name(sheet_name: str) -> str:
    name = sheet_name.strip()
    return "Total CCAA" if name.lower().startswith("total cc") else name


def latest_month_row(rows):
    candidates = [row for row in rows if str(row[0] or "").strip().lower() in MONTHS]
    if not candidates:
        raise ValueError("No monthly cumulative row found")
    return candidates[-1]


def normalize_workbook(input_path: Path, retrieved_at: str | None = None) -> list[dict]:
    workbook_hash = hashlib.sha256(input_path.read_bytes()).hexdigest()
    retrieved = retrieved_at or datetime.now(timezone.utc).isoformat()
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    records = []
    for sheet_name in workbook.sheetnames:
        row = latest_month_row(workbook[sheet_name].iter_rows(values_only=True))
        records.append({
            "territory": normalize_name(sheet_name),
            "territory_level": "autonomous_community",
            "period": f"2026-{str(MONTHS.index(str(row[0]).strip().lower()) + 1).zfill(2)}",
            "period_label": str(row[0]).strip(),
            "data_status": "advance",
            "unit": "thousands_eur",
            "recognized_revenue_current": row[1],
            "recognized_revenue_capital": row[2],
            "recognized_revenue_non_financial": row[3],
            "recognized_expense_current": row[4],
            "recognized_expense_capital": row[5],
            "recognized_expense_non_financial": row[6],
            "source_url": SOURCE_URL,
            "retrieved_at": retrieved,
            "raw_workbook_sha256": workbook_hash,
            "dataset_version": "cimcanet-2026-05",
        })
    return records


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    records = normalize_workbook(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("".join(json.dumps(record, ensure_ascii=False) + "\n" for record in records), encoding="utf-8")
    print(f"normalized={len(records)} sha256={records[0]['raw_workbook_sha256']}")


if __name__ == "__main__":
    main()
